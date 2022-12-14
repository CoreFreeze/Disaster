import sys
from ui_main import *
import datetime
from pytimekr import pytimekr
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.styles.fonts import Font
from emit import setMessageSig
from communication import *
import ui_IPSettings
import socket
import ui_confirm

class myApplication(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        
        self.initCommunication()

        self.initDB()

        self.initUI()
        self.initParameters()

#######################################################################################
#                                    init Func                                        #
#######################################################################################
    def initDB(self):
        self.communication.requestDBFunc(client=None, context=0, request=21, data=None)

    def initCommunication(self):
        bridge = setMessageSig()
        self.communication = ThreadedTCPServer((socket.gethostbyname(socket.gethostname()), int(9080)), chatTCPHandler)
        #self.communication = ThreadedTCPServer(('127.0.0.1', int(9080)), chatTCPHandler)
        self.communication.bridge = bridge
        bridge.DBmessageSig.connect(self.receiveDBData)
        bridge.socketStateSig.connect(self.alertSocketState)
        thread = threading.Thread(target=self.communication.serve_forever)
        thread.start()

    def initParameters(self):
        self.tagListIndex = -1
        self.modifyDisastorType = -1
        today = str(datetime.date.today())
        self.ui.reportDate.setDate(QDate.fromString(today, 'yyyy-MM-dd'))
        self.ui.tagStartDate.setDate(QDate.fromString(today, 'yyyy-MM-dd'))
        self.ui.tagEndDate.setDate(QDate.fromString(today, 'yyyy-MM-dd'))
        self.ui.dataSearchEdit.setDate(QDate.fromString(today, 'yyyy-MM-dd'))

    def initUI(self):
        self.ui.dataTable.setFocus()
        self.ui.disasterInputButton.clicked.connect(self.disasterAdd)
        self.ui.spotInputButton.clicked.connect(self.spotAdd)
        self.ui.programInputButton.clicked.connect(self.programAdd)
        self.ui.tagTable.setColumnCount(3)

        self.communication.requestDBFunc(client=None, context=1, request=1,data=None)
        
        self.communication.requestDBFunc(client=None, context=2, request=13,data=None)
        
        self.ui.dataSearchButton.clicked.connect(self.searchData)
        self.ui.nextDate.clicked.connect(self.nextDate)
        self.ui.preDate.clicked.connect(self.preDate)
        
        self.ui.tagTable.itemDoubleClicked.connect(self.tagModify)

        self.ui.tagAddButton.clicked.connect(self.tagAdd)
        self.ui.tagDelButton.clicked.connect(self.tagDel)

        self.ui.findTaggedProgram.clicked.connect(self.searchProgramName)
        self.ui.programNameAddButton.clicked.connect(self.programNameAdd)
        self.ui.programNameTable.itemDoubleClicked.connect(self.programModify)
        self.ui.programNameDelButton.clicked.connect(self.programDel)

        self.ui.spotAddButton.clicked.connect(self.spotNameAdd)
        self.ui.spotDelButton.clicked.connect(self.spotNameDel)
        self.ui.spotTable.itemDoubleClicked.connect(self.spotNameModify)

        self.ui.reportDate.dateChanged.connect(lambda: self.communication.requestDBFunc(client=None, context=3,request=3,data=(self.ui.reportDate.date().toString('yyyy-MM-dd'),)))
        
        self.ui.createTagSearchButton.clicked.connect(self.searchCreateProgramName)
        self.ui.createFileButton.clicked.connect(self.createFile)
        
        self.ui.dataTable.itemDoubleClicked.connect(self.findDisasterToModify)
        
        self.ui.delDisasterButton.clicked.connect(self.deleteDisaster)

        self.ui.tagModifyButton.clicked.connect(self.modifyTag)
        self.ui.programModifyButton.clicked.connect(self.modifyProgram)
        self.ui.spotModifyButton.clicked.connect(self.modifySpot)

        self.ui.reportStartTime.timeChanged.connect(self.updateEndTime)

        self.ui.modifyDisasterButton.clicked.connect(self.modifyDisaster)

        self.ui.filePathButton.clicked.connect(self.browse)

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_F2:
            settingWindow = Settings()
            settingWindow.exec()

    
    def closeEvent(self, e):
        self.communication.shutdown()
        e.accept()

#######################################################################################
#                                Communication Func                                   #
#######################################################################################

    @Slot(str)
    def alertSocketState(self, err):
        msgBox = QMessageBox.critical(self, '통신 에러!', err)
    
    @Slot(int, list)
    def receiveDBData(self, requestContext, recData):
        if requestContext == 0:
            # CreateDB
            pass
        elif requestContext == 1:
            #searchTags
            self.searchTags(recData)
        elif requestContext == 2:
            self.searchSpotNameList(recData)
        elif requestContext == 3:
            self.searchTagToUpdateProgramList(recData)
        elif requestContext == 4:
            # searchData, nextDate, preDate
            self.addToMainTable(recData[1:])
        elif requestContext == 5:
            # disastor add
            self.addToMainTable(recData[1:])
        elif requestContext == 6:
            self.updateProgramList(recData)
        elif requestContext == 7:
            self.addProgramNameList(recData)
        elif requestContext == 9:
            self.communication.requestDBFunc(client=None, context=2, request=13,data=None)
        elif requestContext == 10:
            self.searchedCreateProgramName(recData)
        elif requestContext == 11:
            self.createSheetOfDay(recData[0], recData[1:])
        elif requestContext == 12:
            self.createSheetOfDay(recData[0], recData[1:])
            self.makeCover(recData[0])
        elif requestContext == 13:
            self.disasterModifyState = recData[0]
            self.modifyDisasterLoad()


#######################################################################################
#                                   Search Func                                       #
#######################################################################################

    def searchData(self):
        date = self.ui.dataSearchEdit.date()
        self.communication.requestDBFunc(client=None, context=4, request=18, data=(date.toString('yyyy-MM-dd'),))

    def nextDate(self):
        date = self.ui.dataSearchEdit.date().addDays(1)
        self.ui.dataSearchEdit.setDate(date)
        self.communication.requestDBFunc(client=None, context=4, request=18, data=(date.toString('yyyy-MM-dd'),))

    def preDate(self):
        date = self.ui.dataSearchEdit.date().addDays(-1)
        self.ui.dataSearchEdit.setDate(date)
        self.communication.requestDBFunc(client=None, context=4, request=18, data=(date.toString('yyyy-MM-dd'),))

    def addToMainTable(self, disasterList):
        self.ui.dataTable.clearContents()
        self.ui.dataTable.setRowCount(len(disasterList))
        for index, disaster in enumerate(disasterList):
            self.ui.dataTable.setItem(index, 0, QTableWidgetItem(disaster[0])) # 프로그램 명
            self.ui.dataTable.setItem(index, 1, QTableWidgetItem(disaster[1])) # 재난유형
            self.ui.dataTable.setItem(index, 2, QTableWidgetItem(disaster[2])) # 세부내용
            self.ui.dataTable.setItem(index, 3, QTableWidgetItem(disaster[3])) # 출연자명
            self.ui.dataTable.setItem(index, 4, QTableWidgetItem(disaster[4]+'~'+disaster[5]))# 방송시간
            if (QTime().fromString(disaster[4],'hh:mm:ss').secsTo(QTime().fromString(disaster[5],'hh:mm:ss'))) % 60 == 0:
                duration = (QTime().fromString(disaster[4],'hh:mm:ss').secsTo(QTime().fromString(disaster[5],'hh:mm:ss')))//60
            else:
                duration = (QTime().fromString(disaster[4],'hh:mm:ss').secsTo(QTime().fromString(disaster[5],'hh:mm:ss')))//60 + 1
            self.ui.dataTable.setItem(index, 5,QTableWidgetItem(str(duration)))
        self.disasterModifyState = -1

#######################################################################################
#                                       Tab#1                                         #
#######################################################################################

    def searchTagToUpdateProgramList(self, recData):
        self.ui.reportProgram.clear()

        if recData is None:
            return
        isWeekend = datetime.date.fromisoformat(recData[0]).isoweekday()
        if isWeekend == 6 or isWeekend ==7:
            self.communication.requestDBFunc(client=None, context=6,request=8,data=(recData[1],))
        else:
            self.communication.requestDBFunc(client=None, context=6,request=7,data=(recData[1],))
        
    def updateProgramList(self, recData):
        programNameList = list(map(convertToList, recData))
        self.ui.reportProgram.addItems(programNameList)
        self.ui.reportProgram.setCurrentIndex(-1)

    def updateEndTime(self):
        self.ui.reportEndTime.setTime(self.ui.reportStartTime.time())

#######################################################################################
#                                    Clear Func                                       #
#######################################################################################

    def clearDisasterArea(self):
        self.ui.disasterType.clear()
        self.ui.disasterDetail.clear()

    def clearSpotArea(self):
        self.ui.spotDetail.setCurrentIndex(-1)

    def clearProgramArea(self):
        self.ui.programDetail.clear()
        self.ui.programSpeaker.clear()
        self.ui.programType.clear()

#######################################################################################
#                                      Add Func                                       #
#######################################################################################

    def disasterAdd(self):
        confirmWindow = Confirm()
        returnData = confirmWindow.exec()
        if returnData == 0:
            return

        date = self.ui.reportDate.date()
        programName = self.ui.reportProgram.currentText()
        type = self.ui.disasterType.text()
        detail = self.ui.disasterDetail.text()
        startTime = self.ui.reportStartTime.time()
        endTime = self.ui.reportEndTime.time()

        if type == '' or detail == '':
            return

        self.communication.requestDBFunc(client=None, context=5, request=20, data=(date.toString('yyyy-MM-dd'), programName, type, detail, '재난수신클라이언트', startTime.toString('hh:mm:ss'), endTime.toString('hh:mm:ss')))
        self.ui.dataSearchEdit.setDate(date)
        self.clearDisasterArea()

    def spotAdd(self):
        confirmWindow = Confirm()
        returnData = confirmWindow.exec()
        if returnData == 0:
            return
        date = self.ui.reportDate.date()
        programName = self.ui.reportProgram.currentText()
        detail = self.ui.spotDetail.currentText()
        startTime = self.ui.reportStartTime.time()
        endTime = self.ui.reportEndTime.time()

        if detail == '':
            return

        self.communication.requestDBFunc(client=None, context=5, request=20, data=(date.toString('yyyy-MM-dd'), programName, None, detail, 'SPOT', startTime.toString('hh:mm:ss'), endTime.toString('hh:mm:ss')))
        self.ui.dataSearchEdit.setDate(date)
        self.clearSpotArea()

    def programAdd(self):
        confirmWindow = Confirm()
        returnData = confirmWindow.exec()
        if returnData == 0:
            return
        date = self.ui.reportDate.date()
        programName = self.ui.reportProgram.currentText()
        type = self.ui.programType.text()
        detail = self.ui.programDetail.text()
        speaker = self.ui.programSpeaker.text()
        startTime = self.ui.reportStartTime.time()
        endTime = self.ui.reportEndTime.time()

        if type == '' or detail == '':
            return
        if speaker == '':
            speaker = '-'

        self.communication.requestDBFunc(client=None, context=5, request=20, data=(date.toString('yyyy-MM-dd'), programName, type, detail, speaker, startTime.toString('hh:mm:ss'), endTime.toString('hh:mm:ss')))
        self.ui.dataSearchEdit.setDate(date)
        self.clearProgramArea()

#######################################################################################
#                                      Modify Func                                    #
#######################################################################################

    def findDisasterToModify(self):
        self.clearDisasterArea()
        self.clearSpotArea()
        self.clearProgramArea()
        self.ui.spotDetail.setCurrentIndex(-1)
        selectedRowIndex = self.ui.dataTable.currentRow()
        oldStartTime, oldEndTime = self.ui.dataTable.item(selectedRowIndex, 4).text().split('~')
        self.communication.requestDBFunc(client=None, context=13, request=22, data=(self.ui.dataSearchEdit.date().toString('yyyy-MM-dd'), oldStartTime, oldEndTime))

    def modifyDisasterLoad(self):
        selectedRowIndex = self.ui.dataTable.currentRow()
        programName = self.ui.dataTable.item(selectedRowIndex, 0).text()
        self.ui.reportDate.setDate(self.ui.dataSearchEdit.date())
        self.ui.reportProgram.setCurrentIndex(self.ui.reportProgram.findText(programName))
        oldStartTime, oldEndTime = self.ui.dataTable.item(selectedRowIndex, 4).text().split('~')
        self.ui.reportStartTime.setTime(QTime().fromString(oldStartTime,"hh:mm:ss"))
        self.ui.reportEndTime.setTime(QTime().fromString(oldEndTime,"hh:mm:ss"))
        if self.ui.dataTable.item(selectedRowIndex, 3).text() == 'SPOT':
            # 스팟
            self.ui.spotDetail.setCurrentIndex(self.ui.spotDetail.findText(self.ui.dataTable.item(selectedRowIndex, 2).text()))
            self.modifyDisastorType = 1
        elif self.ui.dataTable.item(selectedRowIndex, 3).text() == '재난수신클라이언트':
            # 재난
            self.ui.disasterDetail.setText(self.ui.dataTable.item(selectedRowIndex, 2).text())
            self.ui.disasterType.setText(self.ui.dataTable.item(selectedRowIndex, 1).text())
            self.modifyDisastorType = 2
        else:
            self.ui.programDetail.setText(self.ui.dataTable.item(selectedRowIndex, 2).text())
            self.ui.programType.setText(self.ui.dataTable.item(selectedRowIndex, 1).text())
            self.ui.programSpeaker.setText(self.ui.dataTable.item(selectedRowIndex, 3).text())
            self.modifyDisastorType = 3

    def modifyDisaster(self):
        confirmWindow = Confirm()
        confirmWindow.setValue('수정하시겠습니까?')
        returnData = confirmWindow.exec()
        if returnData == 0:
            return

        date = self.ui.reportDate.date()
        programName = self.ui.reportProgram.currentText()
        startTime = self.ui.reportStartTime.time()
        endTime = self.ui.reportEndTime.time()
        
        if self.modifyDisastorType == 2:
            #재난
            inputType = self.ui.disasterType.text()
            detail = self.ui.disasterDetail.text()
            self.communication.requestDBFunc(client=None, context=5, request=23, data=(self.disasterModifyState, date.toString('yyyy-MM-dd'), programName, inputType, detail, '재난수신클라이언트', startTime.toString('hh:mm:ss'), endTime.toString('hh:mm:ss')))
        elif self.modifyDisastorType == 1:
            #SPOT
            detail = self.ui.spotDetail.currentText()
            self.communication.requestDBFunc(client=None, context=5, request=23, data=(self.disasterModifyState, date.toString('yyyy-MM-dd'), programName, '', detail, 'SPOT', startTime.toString('hh:mm:ss'), endTime.toString('hh:mm:ss')))
        elif self.modifyDisastorType == 3:
            #멘트
            inputType = self.ui.programType.text()
            detail = self.ui.programDetail.text()
            speaker = self.ui.programSpeaker.text()
            self.communication.requestDBFunc(client=None, context=5, request=23, data=(self.disasterModifyState, date.toString('yyyy-MM-dd'), programName, inputType, detail, speaker, startTime.toString('hh:mm:ss'), endTime.toString('hh:mm:ss')))
        self.modifyDisastorType = -1
        self.disasterModifyState = -1

#######################################################################################
#                                      Delete Func                                    #
#######################################################################################

    def deleteDisaster(self):
        self.communication.requestDBFunc(client=None, context=5, request=24, data=(self.disasterModifyState, self.ui.dataSearchEdit.date().toString('yyyy-MM-dd')))

#######################################################################################
#                                       Tab#2                                         #
#######################################################################################

#######################################################################################
#                                       Tag Area                                      #
#######################################################################################

    def tagAdd(self):
        newTag = self.ui.tagInput.text()
        if newTag == '':
            # 입력이 없을 경우
            return
        startDate = self.ui.tagStartDate.date()
        endDate = self.ui.tagEndDate.date()
        if startDate > endDate:
            # 기간이 이상할 경우 
            return
        startDate = startDate.toString('yyyy-MM-dd')
        endDate = endDate.toString('yyyy-MM-dd')
        self.communication.requestDBFunc(client=None, context=1, request=0,data=(newTag,startDate,endDate))

    def tagDel(self):
        if self.tagModifyState < 0:
            # 선택되지 않았을 경우
            return
        self.communication.requestDBFunc(client=None, context=1, request=5,data=(self.ui.tagTable.item(self.tagModifyState, 0).text(),))

    def tagModify(self):
        selectedRowIndex = self.ui.tagTable.currentRow()
        self.ui.tagInput.setText(self.ui.tagTable.item(selectedRowIndex, 0).text())
        self.ui.tagStartDate.setDate(QDate.fromString(self.ui.tagTable.item(selectedRowIndex, 1).text(), "yyyy-MM-dd"))
        self.ui.tagEndDate.setDate(QDate.fromString(self.ui.tagTable.item(selectedRowIndex, 2).text(), "yyyy-MM-dd"))
        self.tagModifyState = selectedRowIndex

    def modifyTag(self):
        newTag = self.ui.tagInput.text()
        if newTag == '':
            # 입력이 없을 경우
            return
        startDate = self.ui.tagStartDate.date()
        endDate = self.ui.tagEndDate.date()
        if startDate > endDate:
            # 기간이 이상할 경우 
            return
        startDate = startDate.toString('yyyy-MM-dd')
        endDate = endDate.toString('yyyy-MM-dd')
        self.communication.requestDBFunc(client=None, context=1, request=4,data=(newTag,startDate,endDate, self.ui.tagTable.item(self.tagModifyState, 0).text()))
        


#######################################################################################
#                                  Tag&Program Area                                   #
#######################################################################################

    def searchTags(self, recData):
        self.ui.tagInput.clear()
        self.ui.tagList.clear()
        self.ui.createTagNameList.clear()
        self.ui.tagTable.setSortingEnabled(False)
        self.ui.tagTable.setRowCount(len(recData))
        for rowIndex, tag in enumerate(recData):
            for columnIndex, data in enumerate(tag):
                if columnIndex == 0:
                    continue
                self.ui.tagTable.setItem(rowIndex, columnIndex-1,QTableWidgetItem(data))
                if columnIndex == 1:
                    self.ui.tagList.addItem(data)
                    self.ui.createTagNameList.addItem(data)
        self.ui.tagTable.setSortingEnabled(True)
        self.tagModifyState = -1
        self.ui.tagTable.setCurrentCell(-1,1,QItemSelectionModel.Clear)
        self.ui.tagList.setCurrentIndex(-1)
        self.ui.createTagNameList.setCurrentIndex(-1)

#######################################################################################
#                                    Program Area                                     #
#######################################################################################

    def searchProgramName(self):
        self.tagListIndex = self.ui.tagList.currentIndex()
        if self.tagListIndex < 0:
            return
        self.ui.programNameTable.clearContents()
        self.ui.programNameInput.clear()
        self.ui.locationBox.setCheckState(Qt.CheckState.Unchecked)
        self.ui.weekendBox.setCheckState(Qt.CheckState.Unchecked)
        tagName = self.ui.tagList.itemText(self.tagListIndex)
        self.communication.requestDBFunc(client=None, context=7, request=6,data=(tagName,))
        
    def addProgramNameList(self,recData):
        self.ui.programNameTable.setRowCount(len(recData))
        for index, row in enumerate(recData):
            self.ui.programNameTable.setItem(index, 0, QTableWidgetItem(row[0]))
            if row[3] == 1:
                self.ui.programNameTable.setItem(index, 1, QTableWidgetItem('주말'))
            if row[4] == 1:
                self.ui.programNameTable.setItem(index, 2, QTableWidgetItem('로컬'))
        self.ui.programNameTable.setCurrentCell(-1,1,QItemSelectionModel.Clear)
        self.programModifyState = -1

    def programNameAdd(self):
        if self.tagListIndex < 0:
            return
        programName = self.ui.programNameInput.text()
        tagName = self.ui.tagList.itemText(self.tagListIndex)
        location = self.ui.locationBox.checkState()
        isWeekend = self.ui.weekendBox.checkState()
        
        if location == Qt.CheckState.Checked:
            location = True
        else:
            location = False
        if isWeekend == Qt.CheckState.Checked:
            isWeekend = True
        else:
            isWeekend = False


        self.communication.requestDBFunc(client=None, context=8, request=10, data=(tagName, programName, location, isWeekend))
        
        self.ui.tagList.setCurrentIndex(self.tagListIndex)
        self.searchProgramName()

    def modifyProgram(self):

        if self.programModifyState < 0:
            return

        programName = self.ui.programNameInput.text()
        tagName = self.ui.tagList.itemText(self.tagListIndex)
        location = self.ui.locationBox.checkState()
        isWeekend = self.ui.weekendBox.checkState()

        if location is Qt.CheckState.Checked:
            location = True
        else:
            location = False

        if isWeekend is Qt.CheckState.Checked:
            isWeekend = True
        else:
            isWeekend = False

        oldIsWeekend = self.ui.programNameTable.item(self.programModifyState, 1)
        if oldIsWeekend is None:
            oldIsWeekend = 0
        else:
            oldIsWeekend = 1
        
        self.communication.requestDBFunc(client=None, context=8, request=11, data=(tagName, programName, self.ui.programNameTable.item(self.programModifyState, 0).text(), oldIsWeekend, location, isWeekend))
        self.ui.tagList.setCurrentIndex(self.tagListIndex)
        self.searchProgramName()

    def programModify(self):
        selectedRowIndex = self.ui.programNameTable.currentRow()
        self.ui.weekendBox.setChecked(Qt.CheckState.Unchecked)
        self.ui.locationBox.setChecked(Qt.CheckState.Unchecked)
        self.ui.programNameInput.setText(self.ui.programNameTable.item(selectedRowIndex, 0).text())
        if self.ui.programNameTable.item(selectedRowIndex, 1) is not None:
            self.ui.weekendBox.setChecked(Qt.CheckState.Checked)
        if self.ui.programNameTable.item(selectedRowIndex, 2) is not None:
            self.ui.locationBox.setChecked(Qt.CheckState.Checked)
        self.programModifyState = selectedRowIndex

    def programDel(self):
        if self.programModifyState < 0:
            # 선택되지 않았을 경우
            return
        tagName = self.ui.tagList.itemText(self.tagListIndex)
        oldIsWeekend = self.ui.programNameTable.item(self.programModifyState, 1)
        if oldIsWeekend is None:
            oldIsWeekend = 0
        else:
            oldIsWeekend = 1
        self.communication.requestDBFunc(client=None, context=8, request=12, data=(tagName, self.ui.programNameTable.item(self.programModifyState, 0).text(), oldIsWeekend))
        self.searchProgramName()

#######################################################################################
#                                      SPOT Area                                      #
#######################################################################################

    def spotNameAdd(self):
        newSpot = self.ui.spotInput.text()
        newSpotType = self.ui.spotTypeInput.text()
        if newSpot == '':
            # 입력이 없을 경우
            return
        self.communication.requestDBFunc(client=None, context=2, request=15,data=(newSpot,newSpotType))
        
    def spotNameDel(self):
        if self.spotModifyState < 0:
            # 선택되지 않았을 경우
            return
        self.communication.requestDBFunc(client=None, context=2, request=17, data=(self.ui.spotTable.item(self.spotModifyState, 0).text(),))

    def modifySpot(self):
        newSpot = self.ui.spotInput.text()
        newSpotType = self.ui.spotTypeInput.text()
        if newSpot == '':
            # 입력이 없을 경우
            return
        self.communication.requestDBFunc(client=None, context=2, request=16,data=(newSpot, self.ui.spotTable.item(self.spotModifyState, 0).text(), newSpotType))

    def spotNameModify(self):
        selectedRowIndex = self.ui.spotTable.currentRow()
        self.ui.spotInput.setText(self.ui.spotTable.item(selectedRowIndex, 0).text())
        self.ui.spotTypeInput.setText(self.ui.spotTable.item(selectedRowIndex, 1).text())
        self.spotModifyState = selectedRowIndex

    def searchSpotNameList(self, recData):
        self.ui.spotTable.clearContents()
        self.ui.spotInput.clear()
        self.ui.spotTypeInput.clear()
        self.ui.spotDetail.clear()
        self.ui.spotTable.setRowCount(len(recData))
        for index, row in enumerate(recData):
            self.ui.spotTable.setItem(index, 0, QTableWidgetItem(row[0]))
            self.ui.spotTable.setItem(index, 1, QTableWidgetItem(row[1]))
        spotList = list(map(convertToList,recData))
        self.ui.spotDetail.addItems(spotList)
        self.ui.spotDetail.setCurrentIndex(-1)
        self.ui.spotTable.setCurrentCell(-1,1,QItemSelectionModel.Clear)
        self.spotModifyState = -1

#######################################################################################
#                                        Tab#3                                        #
#######################################################################################

    def browse(self):
        directory = QFileDialog.getExistingDirectory(self, "저장 경로",
                QDir.currentPath())

        if directory:
            self.ui.filePath.setText(directory)

    def searchCreateProgramName(self):
        self.tagListIndex = self.ui.createTagNameList.currentIndex()
        if self.tagListIndex < 0:
            return
        self.ui.createProgramNameList.clear()
        tagName = self.ui.createTagNameList.itemText(self.tagListIndex)
        self.communication.requestDBFunc(client=None, context=10, request=6,data=(tagName,))

    def searchedCreateProgramName(self, recData):
        if len(recData)<=0:
            return
        programNameList = ['통합']
        programNameList = programNameList + list(map(convertToList,recData))
        self.ui.createProgramNameList.addItems(programNameList)
        self.ui.createProgramNameList.setCurrentIndex(-1)
        self.ui.createStartDate.setMinimumDate(QDate().fromString(recData[0][1], 'yyyy-MM-dd'))
        self.ui.createEndDate.setMinimumDate(QDate().fromString(recData[0][1], 'yyyy-MM-dd'))
        self.ui.createStartDate.setMaximumDate(QDate().fromString(recData[0][2], 'yyyy-MM-dd'))
        self.ui.createEndDate.setMaximumDate(QDate().fromString(recData[0][2], 'yyyy-MM-dd'))
        self.ui.createEndDate.setDate(QDate().fromString(recData[0][2], 'yyyy-MM-dd'))

    def createFile(self):
        isTotal = False
        startDate=self.ui.createStartDate.date()
        endDate=self.ui.createEndDate.date()
        self.myExcelWorkbook = Workbook()
        programName = self.ui.createProgramNameList.currentText()
        if programName == '':
            return
        elif programName == '통합':
            isTotal = True

        self.disasterWeekendMinCount = {}
        self.disasterWeekdayMinCount = {}

        for index in range(0,startDate.daysTo(endDate)+1):
            # DB 요청 정보 받은 후 새로 파일을 만들지 말지 체크해서 context 다르게 보내기
            currentDate = startDate.addDays(index)
            currentMonth = currentDate.month()
            nextMonth = currentDate.addDays(1).month()
            context = 11
            if currentMonth != nextMonth or index == startDate.daysTo(endDate):
                context = 12
            if isTotal:
                self.communication.requestDBFunc(client=None, context=context, request=18, data=(currentDate.toString('yyyy-MM-dd'),))
            else:
                self.communication.requestDBFunc(client=None, context=context, request=25, data=(currentDate.toString('yyyy-MM-dd'), programName))
            

    def checkHoliday(self, currentDate):
        retData = 0
        currnetPythonDate = currentDate.toPython()
        KORholidays = pytimekr.holidays(year=currentDate.year())
        subHolidays = pytimekr.red_days(pytimekr.lunar_newyear(year=currentDate.year())) + pytimekr.red_days(pytimekr.chuseok(year=currentDate.year())) + [pytimekr.children(year=currentDate.year()), pytimekr.independence(year=currentDate.year()), pytimekr.hangul(year=currentDate.year()), pytimekr.samiljeol(year=currentDate.year()), pytimekr.foundation(year=currentDate.year())]
        
        if currnetPythonDate in KORholidays:
            retData = 1 # 빨간색
        else:
            dayNumber = currnetPythonDate.isoweekday()
            if dayNumber == 7:
                retData = 1 # 빨간색
            elif dayNumber == 6:
                retData = 2 # 파란색
            elif dayNumber == 1: # 월요일의 경우
                preDate = currentDate.addDays(-1).toPython()
                prePreDate = currentDate.addDays(-2).toPython()
                if preDate in subHolidays or prePreDate in subHolidays:
                    retData = 1 # 대체공휴일
            else:
                retData = False
        return retData


    def createSheetOfDay(self, currentDate, rows):
        currentDate = QDate().fromString(currentDate,'yyyy-MM-dd')

        totalOfDay = {}

        ws = self.myExcelWorkbook.create_sheet(currentDate.toString('MMdd'))
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 80
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['J'].width = 15

        dayInfo = self.checkHoliday(currentDate)
        
        # 토요일일 경우
        if dayInfo == 2:
            ws.sheet_properties.tabColor = "0000FF"
        # 일요일, 공휴일, 대체공휴일의 경우
        elif dayInfo == 1:
            ws.sheet_properties.tabColor = "FF0000"

        row1 = ['방송', '일자', '프로그램', '재난유형','세부내용','출연자명','방송시간','시간(분)', '', '프로그램', '시간']
        ws.append(row1)
        for i in range(1, 12):
            if i == 9:
                continue
            ws.cell(row=1, column=i).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=i).fill = PatternFill(fill_type='solid',start_color='00CCFFCC',end_color='00CCFFCC')
            if i < 9:
                ws.cell(row=1, column=i).border = Border(left=Side(border_style='thin',color='00000000'),right=Side(border_style='thin',color='00000000'),top=Side(border_style='thin',color='00000000'),bottom=Side(border_style='thin',color='00000000'))
                ws.cell(row=2, column=i).border = Border(left=Side(border_style='thin',color='00000000'),right=Side(border_style='thin',color='00000000'),top=Side(border_style='thin',color='00000000'),bottom=Side(border_style='thin',color='00000000'))


        maxRowIndex = 2
        preProgramName = ''
        preRowIndex = 1
        ws.cell(row=2, column=1, value='경북')
        for rowIndex, row in enumerate(rows):
            localProgram = 0
            ws.cell(row=rowIndex+2, column=1).border = Border(left=Side(border_style='thin',color='00000000'),right=Side(border_style='thin',color='00000000'),top=Side(border_style='thin',color='00000000'),bottom=Side(border_style='thin',color='00000000'))

            ws.cell(row=rowIndex+2, column=3, value=row[0]) # 프로그램
            ws.cell(row=rowIndex+2, column=3).border = Border(left=Side(border_style='thin',color='00000000'),right=Side(border_style='thin',color='00000000'),top=Side(border_style='thin',color='00000000'),bottom=Side(border_style='thin',color='00000000'))
            ws.cell(row=rowIndex+2, column=3).alignment = Alignment(horizontal='center', vertical='center')
                
            if row[6] == 1:
                ws.cell(row=rowIndex+2, column=3).fill = PatternFill(fill_type='solid', start_color='00FFFF99', end_color='00FFFF99')
                localProgram = 1
            if preProgramName != row[0]:
                ws.merge_cells(start_row=preRowIndex, end_row=rowIndex+1, start_column=3, end_column=3)
                preRowIndex = rowIndex + 2
                preProgramName = row[0]

            for i in range(4, 7):
                ws.cell(row=rowIndex+2, column=i, value=row[i-3]) # 재난유형
                ws.cell(row=rowIndex+2, column=i).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=rowIndex+2, column=i).border = Border(left=Side(border_style='thin',color='00000000'),right=Side(border_style='thin',color='00000000'),top=Side(border_style='thin',color='00000000'),bottom=Side(border_style='thin',color='00000000'))
                    

            ws.cell(row=rowIndex+2, column=7, value=row[4]+'~'+row[5]) # 방송시간
            ws.cell(row=rowIndex+2, column=7).border = Border(left=Side(border_style='thin',color='00000000'),right=Side(border_style='thin',color='00000000'),top=Side(border_style='thin',color='00000000'),bottom=Side(border_style='thin',color='00000000'))
            ws.cell(row=rowIndex+2, column=7).alignment = Alignment(horizontal='center', vertical='center')
            if (QTime().fromString(row[4],'hh:mm:ss').secsTo(QTime().fromString(row[5],'hh:mm:ss'))) % 60 == 0:
                duration = (QTime().fromString(row[4],'hh:mm:ss').secsTo(QTime().fromString(row[5],'hh:mm:ss')))//60
            else:
                duration = (QTime().fromString(row[4],'hh:mm:ss').secsTo(QTime().fromString(row[5],'hh:mm:ss')))//60 + 1
            # 분
            ws.cell(row=rowIndex+2, column=8, value=duration)
            ws.cell(row=rowIndex+2, column=8).border = Border(left=Side(border_style='thin',color='00000000'),right=Side(border_style='thin',color='00000000'),top=Side(border_style='thin',color='00000000'),bottom=Side(border_style='thin',color='00000000'))
            ws.cell(row=rowIndex+2, column=8).alignment = Alignment(horizontal='center', vertical='center')

            dayNum = currentDate.toPython().isoweekday()
            if dayNum == 6 or dayNum == 7:
                if row[0] not in self.disasterWeekendMinCount:
                    self.disasterWeekendMinCount[row[0]] = [localProgram, duration]
                    totalOfDay[row[0]] = [localProgram, duration]
                else:
                    self.disasterWeekendMinCount[row[0]] = [self.disasterWeekendMinCount[row[0]][0], self.disasterWeekendMinCount[row[0]][1]+duration]
                    totalOfDay[row[0]] = [totalOfDay[row[0]][0], totalOfDay[row[0]][1]+duration]
            else:
                if row[0] not in self.disasterWeekdayMinCount:
                    self.disasterWeekdayMinCount[row[0]] = [localProgram,duration]
                    totalOfDay[row[0]] = [localProgram, duration]
                else:
                    self.disasterWeekdayMinCount[row[0]] = [self.disasterWeekdayMinCount[row[0]][0], self.disasterWeekdayMinCount[row[0]][1]+duration]
                    totalOfDay[row[0]] = [totalOfDay[row[0]][0], totalOfDay[row[0]][1]+duration]

            maxRowIndex = rowIndex + 2

        if maxRowIndex > 2:
            ws.merge_cells(start_row=preRowIndex, end_row=maxRowIndex, start_column=3, end_column=3)
            
        ws.merge_cells(start_row=2, end_row=maxRowIndex,start_column=1,end_column=1)
        ws.cell(row=2,column=1).alignment = Alignment(horizontal='center', vertical='top')

        ws.cell(row=2,column=2).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=2, column=2, value=currentDate.toString('MM월 dd일'))
        ws.merge_cells(start_row=2, end_row=maxRowIndex,start_column=2,end_column=2)

        totalValue = 0
        for index, programName in enumerate(totalOfDay):
            ws.cell(row=index+2, column=10, value=programName)
            if totalOfDay[programName][0]:
                ws.cell(row=index+2, column=10).fill = PatternFill(fill_type='solid', start_color='00FDE9D9', end_color='00FDE9D9')
            ws.cell(row=index+2, column=11, value=totalOfDay[programName][1])
            totalValue = totalValue + totalOfDay[programName][1]
        
        ws.cell(row=len(totalOfDay)+3, column=10, value='총합')
        ws.cell(row=len(totalOfDay)+3, column=11, value=totalValue)

    def makeCover(self, currentDateString):
        ws = self.myExcelWorkbook[self.myExcelWorkbook.sheetnames[0]]
        currentDate = datetime.date.fromisoformat(currentDateString)
        title = currentDate.strftime('%y년 %m월 경북방송 재난방송 결과 취합')
        ws.title = title
        ws.merge_cells(start_row=1, end_row=2,start_column=2,end_column=7)
        ws.cell(row=1, column=2, value=title)
        ws.cell(row=1, column=2).font = Font(size=20)
        ws.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=1, column=2).fill = PatternFill(fill_type='solid', start_color='00C6EFCE', end_color='00C6EFCE')
                

        ws.merge_cells(start_row=4, end_row=4,start_column=3,end_column=4)
        ws.cell(row=4, column=3).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=4, column=3, value='주중')
        ws.cell(row=4, column=3).fill = PatternFill(fill_type='solid', start_color='00DAEEF3', end_color='00DAEEF3')
        
        ws.merge_cells(start_row=4, end_row=4,start_column=5,end_column=6)
        ws.cell(row=4, column=5).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=4, column=5, value='주말')
        ws.cell(row=4, column=5).fill = PatternFill(fill_type='solid', start_color='00DAEEF3', end_color='00DAEEF3')

        ws.cell(row=5, column=3, value='프로그램')
        ws.cell(row=5, column=3).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=5, column=3).fill = PatternFill(fill_type='solid', start_color='00EBF1DE', end_color='00EBF1DE')
        ws.cell(row=5, column=4, value='시간(분)')
        ws.cell(row=5, column=4).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=5, column=4).fill = PatternFill(fill_type='solid', start_color='00EBF1DE', end_color='00EBF1DE')
        ws.cell(row=5, column=5, value='프로그램')
        ws.cell(row=5, column=5).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=5, column=5).fill = PatternFill(fill_type='solid', start_color='00EBF1DE', end_color='00EBF1DE')
        ws.cell(row=5, column=6, value='시간(분)')
        ws.cell(row=5, column=6).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=5, column=6).fill = PatternFill(fill_type='solid', start_color='00EBF1DE', end_color='00EBF1DE')
       
        weekDayIndex = 6
        weekDayTotal = 0
        for index, programName in enumerate(self.disasterWeekdayMinCount):
            ws.cell(row=index+6, column=3, value=programName)
            ws.cell(row=index+6, column=4, value=self.disasterWeekdayMinCount[programName][1])
            ws.cell(row=index+6, column=3).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=index+6, column=4).alignment = Alignment(horizontal='center', vertical='center')
            if self.disasterWeekdayMinCount[programName][0] == 1:
                ws.cell(row=index+6, column=3).fill = PatternFill(fill_type='solid', start_color='00FDE9D9', end_color='00FDE9D9')
            weekDayTotal += self.disasterWeekdayMinCount[programName][1]
            weekDayIndex = index + 6

        weekEndIndex = 6
        weekEndTotal = 0
        for index, programName in enumerate(self.disasterWeekendMinCount):
            ws.cell(row=index+6, column=5, value=programName)
            ws.cell(row=index+6, column=6, value=self.disasterWeekendMinCount[programName][1])
            ws.cell(row=index+6, column=5).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=index+6, column=6).alignment = Alignment(horizontal='center', vertical='center')
            if self.disasterWeekendMinCount[programName][0] == 1:
                ws.cell(row=index+6, column=5).fill = PatternFill(fill_type='solid', start_color='00FDE9D9', end_color='00FDE9D9')
            weekEndTotal += self.disasterWeekendMinCount[programName][1]
            weekEndIndex = index + 6
        
        sumContentsIndex = max(weekDayIndex, weekEndIndex) + 1
        ws.cell(row=sumContentsIndex, column=3, value = '주간 총합')
        ws.cell(row=sumContentsIndex, column=3).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=sumContentsIndex, column=3).fill = PatternFill(fill_type='solid', start_color='00FDE9D9', end_color='00FDE9D9')
        ws.cell(row=sumContentsIndex, column=4, value = weekDayTotal)
        ws.cell(row=sumContentsIndex, column=4).alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=sumContentsIndex, column=5, value = '주말 총합')
        ws.cell(row=sumContentsIndex, column=5).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=sumContentsIndex, column=5).fill = PatternFill(fill_type='solid', start_color='00FDE9D9', end_color='00FDE9D9')
        ws.cell(row=sumContentsIndex, column=6, value = weekEndTotal)
        ws.cell(row=sumContentsIndex, column=6).alignment = Alignment(horizontal='center', vertical='center')

        sumContentsIndex += 1
        ws.merge_cells(start_row=sumContentsIndex, end_row=sumContentsIndex, start_column=3, end_column=4)
        ws.cell(row=sumContentsIndex, column=3, value = '총합')
        ws.cell(row=sumContentsIndex, column=3).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=sumContentsIndex, column=3).fill = PatternFill(fill_type='solid', start_color='00D8E4BC', end_color='00D8E4BC')
        ws.merge_cells(start_row=sumContentsIndex, end_row=sumContentsIndex, start_column=5, end_column=6)
        ws.cell(row=sumContentsIndex, column=5, value = weekDayTotal+weekEndTotal)
        ws.cell(row=sumContentsIndex, column=5).alignment = Alignment(horizontal='center', vertical='center')

        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 30

        for i in range(4,sumContentsIndex+1):
            ws.row_dimensions[i].height = 30
        

        self.myExcelWorkbook.save(self.ui.filePath.text()+'/'+title+'.xlsx')
        self.myExcelWorkbook.close()
        self.disasterWeekendMinCount = {}
        self.disasterWeekdayMinCount = {}
        self.myExcelWorkbook = Workbook()


#######################################################################################
#                                     Sub Func                                        #
#######################################################################################

def convertToList(dataRows):
    return(dataRows[0])

#######################################################################################
#                                     Setting                                         #
#######################################################################################

class Settings(QDialog):
    def __init__(self):
        super().__init__()
        self.ui = ui_IPSettings.Ui_Dialog()
        self.ui.setupUi(self)
        self.ui.lineEdit.setText(socket.gethostbyname(socket.gethostname()))
        self.ui.lineEdit.setDisabled(True)
        self.show()


class Confirm(QDialog):
    def __init__(self):
        super().__init__()
        self.ui = ui_confirm.Ui_Dialog()
        self.ui.setupUi(self)
        self.show()

    def setValue(self, stringData):
        self.ui.label.setText(stringData)

#######################################################################################
#                                        Main                                        #
#######################################################################################

if __name__ == '__main__':
    app = QApplication()
    window = myApplication()
    window.show()

    sys.exit(app.exec_())